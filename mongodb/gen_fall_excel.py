"""
gen_fall_excel.py — Tạo file Excel mô phỏng 1 window (100 mẫu @ 50 Hz) sự kiện ngã.
Chạy: python gen_fall_excel.py
"""
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ─── Sinh dữ liệu mô phỏng ────────────────────────────────────────────────────
SAMPLE_RATE = 50          # Hz
N = 100                   # 1 window = 100 mẫu = 2.0 s
dt = 1.0 / SAMPLE_RATE   # 0.02 s
rng = np.random.default_rng(seed=42)

t = np.linspace(0, (N - 1) * dt, N)

def gauss(center, width, amp):
    return amp * np.exp(-0.5 * ((t - center) / width) ** 2)

def exp_decay(start, tau, amp):
    out = np.zeros(N)
    mask = t >= start
    out[mask] = amp * np.exp(-(t[mask] - start) / tau)
    return out

# ── Accelerometer (g) ─────────────────────────────────────────────────────────
ax = np.zeros(N)
ay = np.zeros(N)
az = np.ones(N) * 1.0   # gravity baseline

# Tilting (0.75–0.90 s)
for i, ti in enumerate(t):
    if 0.75 <= ti < 0.90:
        p = (ti - 0.75) / 0.15
        az[i] -= p * 0.6
        ay[i] += p * 0.4

# Impact (~0.975 s)
az += gauss(0.975, 0.022, 11.5)
ax += gauss(0.975, 0.028, -10.5)
ay += gauss(0.995, 0.025, 8.5)

# Oscillation (1.05–1.4 s)
osc = exp_decay(1.05, 0.14, 1.0)
ax += 4.5 * osc * np.sin(2 * np.pi * 8 * (t - 1.05))
ay += 3.5 * osc * np.cos(2 * np.pi * 8 * (t - 1.05))
az += 2.0 * osc * np.sin(2 * np.pi * 6.5 * (t - 1.05))

# Still (1.4–2.0 s)
for i, ti in enumerate(t):
    if ti >= 1.4:
        p = min((ti - 1.4) / 0.2, 1.0)
        ax[i] *= (1 - p) * 0.3
        ay[i] *= (1 - p) * 0.3
        az[i] = az[i] * (1 - p) + 1.0 * p

ax += rng.normal(0, 0.06, N)
ay += rng.normal(0, 0.06, N)
az += rng.normal(0, 0.06, N)
ax = np.clip(ax, -14, 14)
ay = np.clip(ay, -14, 14)
az = np.clip(az, -14, 14)

# ── Gyroscope (°/s) ───────────────────────────────────────────────────────────
gx = np.zeros(N)
gy = np.zeros(N)
gz = np.zeros(N)

# Pre-tilt gyro
for i, ti in enumerate(t):
    if 0.75 <= ti < 0.95:
        p = (ti - 0.75) / 0.20
        gz[i] += p * 80
        gy[i] += p * 50

gx += gauss(0.970, 0.030,  430)
gy += gauss(0.985, 0.030, -480)
gz += gauss(0.975, 0.025,  520)

g_osc = exp_decay(1.08, 0.12, 1.0)
gx += 200 * g_osc * np.sin(2 * np.pi * 6.0 * (t - 1.08))
gy += 240 * g_osc * np.cos(2 * np.pi * 6.0 * (t - 1.08))
gz += 160 * g_osc * np.sin(2 * np.pi * 5.5 * (t - 1.08) + 0.8)

gx += rng.normal(0, 3, N)
gy += rng.normal(0, 3, N)
gz += rng.normal(0, 3, N)
gx = np.clip(gx, -600, 600)
gy = np.clip(gy, -600, 600)
gz = np.clip(gz, -600, 600)

# Tính magnitude
acc_mag  = np.sqrt(ax**2 + ay**2 + az**2)
gyro_mag = np.sqrt(gx**2 + gy**2 + gz**2)

# Phase label
def get_phase(ti):
    if ti < 0.75:   return "Pre-fall"
    if ti < 0.90:   return "Tilting"
    if ti < 0.95:   return "Free-fall"
    if ti < 1.10:   return "Impact"
    if ti < 1.40:   return "Oscillation"
    return "Still"

# ─── Tạo Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fall_IMU_Window"

# ── Màu sắc ───────────────────────────────────────────────────────────────────
CLR = {
    "header_acc":  "1F6FEB",  # xanh dương
    "header_gyro": "3FB950",  # xanh lá
    "header_meta": "6E40C9",  # tím
    "header_mag":  "D29922",  # vàng
    "phase_pre":   "E8F4FD",
    "phase_tilt":  "FFF3CD",
    "phase_ff":    "FDE8E8",
    "phase_impact":"F85149",
    "phase_osc":   "EAD8FF",
    "phase_still": "E8F5E9",
    "white":       "FFFFFF",
    "text_white":  "FFFFFF",
    "text_dark":   "0D1117",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="0D1117", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center():
    return Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color="D0D7DE")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Tiêu đề chính ─────────────────────────────────────────────────────────────
ws.merge_cells("A1:L1")
title_cell = ws["A1"]
title_cell.value = "IMU FALL EVENT SIMULATION — 1 Window (100 samples @ 50 Hz)"
title_cell.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
title_cell.fill = fill("0D1117")
title_cell.alignment = center()
ws.row_dimensions[1].height = 28

# ── Sub-header thông số ───────────────────────────────────────────────────────
ws.merge_cells("A2:L2")
info = ws["A2"]
info.value = ("Sampling: 50 Hz  |  dt = 20 ms  |  Window: 100 samples = 2.0 s  |  "
              "ACC ±8 g (±14g clip)  |  GYR ±2000 °/s  |  "
              "Gate: |acc|>7.5g OR |gyro|>240°/s")
info.font = Font(italic=True, color="8B949E", size=9, name="Calibri")
info.fill = fill("161B22")
info.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Header cột ───────────────────────────────────────────────────────────────
headers = [
    ("Sample #",    "header_meta"),
    ("Time (s)",    "header_meta"),
    ("Phase",       "header_meta"),
    ("ax (g)",      "header_acc"),
    ("ay (g)",      "header_acc"),
    ("az (g)",      "header_acc"),
    ("|acc| (g)",   "header_mag"),
    ("gx (°/s)",    "header_gyro"),
    ("gy (°/s)",    "header_gyro"),
    ("gz (°/s)",    "header_gyro"),
    ("|gyro| (°/s)","header_mag"),
    ("Gate PASS?",  "header_meta"),
]
ws.row_dimensions[3].height = 22
for col_idx, (hdr, clr_key) in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=hdr)
    cell.font = font(bold=True, color="FFFFFF", size=10)
    cell.fill = fill(CLR[clr_key])
    cell.alignment = center()
    cell.border = border

# ── Dữ liệu ───────────────────────────────────────────────────────────────────
phase_fill_map = {
    "Pre-fall":    ("E8F4FD", "1F6FEB"),
    "Tilting":     ("FFF3CD", "7D5800"),
    "Free-fall":   ("FFEDE8", "C0392B"),
    "Impact":      ("FFCCCC", "F85149"),
    "Oscillation": ("F0E6FF", "6E40C9"),
    "Still":       ("E8F5E9", "1A7F37"),
}

for i in range(N):
    row = i + 4
    ti = t[i]
    phase = get_phase(ti)
    gate  = "✓ PASS" if (acc_mag[i] > 7.5 or gyro_mag[i] > 240) else "—"
    bg_hex, fg_hex = phase_fill_map[phase]

    values = [
        i + 1,
        round(float(ti), 4),
        phase,
        round(float(ax[i]), 4),
        round(float(ay[i]), 4),
        round(float(az[i]), 4),
        round(float(acc_mag[i]), 4),
        round(float(gx[i]), 2),
        round(float(gy[i]), 2),
        round(float(gz[i]), 2),
        round(float(gyro_mag[i]), 2),
        gate,
    ]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        # Impact row: nền đỏ + chữ đậm trắng
        if phase == "Impact":
            cell.fill = fill("FFD0D0")
            cell.font = font(bold=(col_idx in (4,5,6,7,8,9,10,11)),
                             color=("C0000C" if col_idx in (4,5,6,7,8,9,10,11) else "0D1117"),
                             size=9)
        else:
            cell.fill = fill(bg_hex)
            cell.font = font(color=fg_hex if col_idx == 3 else "0D1117",
                             bold=(col_idx == 3), size=9)
        cell.alignment = Alignment(horizontal="center" if col_idx in (1,2,3,12) else "right",
                                   vertical="center")
        cell.border = border

        # Gate PASS: màu đỏ nổi bật
        if col_idx == 12 and gate == "✓ PASS":
            cell.font = Font(bold=True, color="F85149", size=9, name="Calibri")

    ws.row_dimensions[row].height = 15

# ── Độ rộng cột ───────────────────────────────────────────────────────────────
col_widths = [9, 9, 14, 10, 10, 10, 11, 10, 10, 10, 13, 11]
for col_idx, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Freeze header
ws.freeze_panes = "A4"

# ─── Sheet 2: Thống kê ────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Statistics")
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 16
ws2.column_dimensions["C"].width = 16

ws2.merge_cells("A1:C1")
ws2["A1"].value = "IMU Fall Window — Statistics"
ws2["A1"].font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
ws2["A1"].fill = fill("0D1117")
ws2["A1"].alignment = center()

stat_headers = ["Metric", "Value", "Unit"]
for col_idx, h in enumerate(stat_headers, start=1):
    c = ws2.cell(row=2, column=col_idx, value=h)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill("1F6FEB")
    c.alignment = center()
    c.border = border

stats = [
    ("Samples",             N,                          "—"),
    ("Sample Rate",         SAMPLE_RATE,               "Hz"),
    ("dt",                  f"{dt*1000:.0f}",          "ms"),
    ("Duration",            f"{N*dt:.2f}",             "s"),
    ("ax peak",             f"{ax.max():.4f}",         "g"),
    ("ax trough",           f"{ax.min():.4f}",         "g"),
    ("ay peak",             f"{ay.max():.4f}",         "g"),
    ("az peak",             f"{az.max():.4f}",         "g"),
    ("|acc| max",           f"{acc_mag.max():.4f}",    "g"),
    ("|acc| mean",          f"{acc_mag.mean():.4f}",   "g"),
    ("gx peak",             f"{gx.max():.2f}",         "°/s"),
    ("gy trough",           f"{gy.min():.2f}",         "°/s"),
    ("gz peak",             f"{gz.max():.2f}",         "°/s"),
    ("|gyro| max",          f"{gyro_mag.max():.2f}",   "°/s"),
    ("|gyro| mean",         f"{gyro_mag.mean():.2f}",  "°/s"),
    ("Gate PASS samples",   int(((acc_mag>7.5)|(gyro_mag>240)).sum()), "samples"),
    ("Acc threshold",       7.5,                        "g"),
    ("Gyro threshold",      240,                        "°/s"),
]

alt = ["F0F6FF", "FFFFFF"]
for r_idx, (metric, val, unit) in enumerate(stats, start=3):
    bg = alt[r_idx % 2]
    for col_idx, v in enumerate([metric, val, unit], start=1):
        c = ws2.cell(row=r_idx, column=col_idx, value=v)
        c.font = font(size=10)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left" if col_idx==1 else "center",
                                vertical="center")
        c.border = border
    ws2.row_dimensions[r_idx].height = 18

# ─── Chart Sheet: Accelerometer ───────────────────────────────────────────────
ws3 = wb.create_sheet("Chart_ACC")
# Copy dữ liệu time + acc để chart tham chiếu
ws3["A1"].value = "Time(s)"
ws3["B1"].value = "ax(g)"
ws3["C1"].value = "ay(g)"
ws3["D1"].value = "az(g)"
for i in range(N):
    ws3.cell(row=i+2, column=1, value=round(float(t[i]),4))
    ws3.cell(row=i+2, column=2, value=round(float(ax[i]),4))
    ws3.cell(row=i+2, column=3, value=round(float(ay[i]),4))
    ws3.cell(row=i+2, column=4, value=round(float(az[i]),4))

chart_acc = LineChart()
chart_acc.title = "Accelerometer — Fall Event (50 Hz)"
chart_acc.style = 10
chart_acc.y_axis.title = "Acceleration (g)"
chart_acc.x_axis.title = "Sample index"
chart_acc.height = 14
chart_acc.width  = 26

for col, name, color in [(2,"ax","1F6FEB"),(3,"ay","FF7B72"),(4,"az","3FB950")]:
    data = Reference(ws3, min_col=col, min_row=1, max_row=N+1)
    chart_acc.add_data(data, titles_from_data=True)
    chart_acc.series[-1].graphicalProperties.line.solidFill = color
    chart_acc.series[-1].graphicalProperties.line.width = 18000

ws3.add_chart(chart_acc, "F1")

# Chart: Gyroscope
ws4 = wb.create_sheet("Chart_GYRO")
ws4["A1"].value = "Time(s)"
ws4["B1"].value = "gx(°/s)"
ws4["C1"].value = "gy(°/s)"
ws4["D1"].value = "gz(°/s)"
for i in range(N):
    ws4.cell(row=i+2, column=1, value=round(float(t[i]),4))
    ws4.cell(row=i+2, column=2, value=round(float(gx[i]),2))
    ws4.cell(row=i+2, column=3, value=round(float(gy[i]),2))
    ws4.cell(row=i+2, column=4, value=round(float(gz[i]),2))

chart_gyro = LineChart()
chart_gyro.title = "Gyroscope — Fall Event (50 Hz)"
chart_gyro.style = 10
chart_gyro.y_axis.title = "Angular Velocity (°/s)"
chart_gyro.x_axis.title = "Sample index"
chart_gyro.height = 14
chart_gyro.width  = 26

for col, name, color in [(2,"gx","1F6FEB"),(3,"gy","FF7B72"),(4,"gz","3FB950")]:
    data = Reference(ws4, min_col=col, min_row=1, max_row=N+1)
    chart_gyro.add_data(data, titles_from_data=True)
    chart_gyro.series[-1].graphicalProperties.line.solidFill = color
    chart_gyro.series[-1].graphicalProperties.line.width = 18000

ws4.add_chart(chart_gyro, "F1")

# ─── Lưu ─────────────────────────────────────────────────────────────────────
OUT = "fall_event_imu_100samples.xlsx"
wb.save(OUT)
print(f"✓ Đã lưu: {OUT}")
print(f"  Sheet 1 : Fall_IMU_Window  — {N} hàng dữ liệu đầy đủ")
print(f"  Sheet 2 : Statistics       — tóm tắt thống kê")
print(f"  Sheet 3 : Chart_ACC        — biểu đồ gia tốc")
print(f"  Sheet 4 : Chart_GYRO       — biểu đồ góc quay")
print(f"\n  |acc| max  = {acc_mag.max():.3f} g   (threshold 7.5g)  → PASS ✓")
print(f"  |gyro| max = {gyro_mag.max():.1f} °/s (threshold 240°/s) → PASS ✓")
print(f"  Gate PASS  = {int(((acc_mag>7.5)|(gyro_mag>240)).sum())} / {N} samples")
